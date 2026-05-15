# -*- coding: utf-8 -*-
"""
===========================================================
  СКРИПТ ЗБІРКИ ДОДАТКУ 4.3
  Синхронізовані пари TSP (TZA) – PM10 (CairCloud)

  Аркуш 1 — Post3_pan2 : TZA_Post3 <-> CairCloud_pan2
  Аркуш 2 — Post7_pan5 : TZA_Post7 <-> CairCloud_pan5

  Вікно синхронізації: ±30 хвилин від строку TZA (7:00 / 19:00)
  Фільтр CairCloud: QC_PM10 = 1
  Фільтр TZA: строки 7 і 19, TSP ≠ None

  ЗАЛЕЖНОСТІ: тільки openpyxl (pip install openpyxl)

  ЗАПУСК:
      python build_Додаток_4_3.py

  Вихідний файл: Додаток_4_3.xlsx  (у тій самій папці)
===========================================================
"""

import os
import sys
from datetime import datetime, timedelta
from zipfile import BadZipFile

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ПОМИЛКА: бібліотека openpyxl не встановлена.")
    print("Виконайте:  pip install openpyxl")
    sys.exit(1)

# ── Шляхи ────────────────────────────────────────────────────────────────────
# Скрипт автоматично визначає свою папку — кладіть його поруч з Додаток_4_3.xlsx
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

BASE    = os.path.join(SCRIPT_DIR, "Додаток_2 Оцифровані дані")
OUT_4_3 = os.path.join(SCRIPT_DIR, "Додаток_4_3.xlsx")
DIG_CC  = os.path.join(BASE, "Processed_CairCloud_DIG")
DIG_TZA = os.path.join(BASE, "Processed_TZA_DIG")

# ── Довідники ─────────────────────────────────────────────────────────────────
MONTHS_UK = ['', 'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 'Червень',
             'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень']

# Повний маппінг ATM-кодів TZA за РД 52.04.186-89
ATM_MAP = {
    0: 'норма',
    3: 'пилова імла',
    4: 'туман',
    5: 'мряка',
    7: 'сніг',
    8: 'дощ',
    9: 'гроза',
}

# ── Читання CairCloud ─────────────────────────────────────────────────────────
def read_caircloud(pan_folder, has_ws=True):
    """
    Завантажити всі записи CairCloud (QC_PM10 = 1).
      pan2: 13 колонок — є WS (індекс 9).
      pan5: 9 колонок  — WS відсутній.
    Повертає: [(datetime, pm10, t_air, rh, ws_or_None), ...]
    """
    CC_PM10, CC_QC, CC_T, CC_RH = 1, 2, 3, 5
    CC_WS = 9 if has_ws else None

    recs = []
    folder_name = os.path.basename(pan_folder)
    for mm in range(1, 13):
        fpath = os.path.join(pan_folder, f"{mm:02d}_{folder_name}.xlsx")
        if not os.path.exists(fpath):
            print(f"    [ПРОПУСК] {os.path.basename(fpath)} — файл відсутній")
            continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except (BadZipFile, Exception) as e:
            print(f"    [ПРОПУСК] {os.path.basename(fpath)} — пошкоджений: {e}")
            continue

        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= CC_QC or row[CC_QC] != 1:
                continue
            try:
                pm10 = float(row[CC_PM10])
                if pm10 < 0:
                    continue
            except (TypeError, ValueError):
                continue

            ts = row[0]
            if isinstance(ts, datetime):
                dt = ts
            elif isinstance(ts, str):
                dt = None
                for fmt in ['%Y-%m-%d %H:%M', '%d.%m.%Y %H:%M', '%Y-%m-%d %H:%M:%S']:
                    try:
                        dt = datetime.strptime(ts.strip(), fmt)
                        break
                    except ValueError:
                        pass
                if dt is None:
                    continue
            else:
                continue

            t_air = rh_v = ws_v = None
            try: t_air = float(row[CC_T])
            except (TypeError, ValueError): pass
            try: rh_v  = float(row[CC_RH])
            except (TypeError, ValueError): pass
            if CC_WS and len(row) > CC_WS:
                try: ws_v = float(row[CC_WS])
                except (TypeError, ValueError): pass

            recs.append((dt, pm10, t_air, rh_v, ws_v))
            count += 1
        wb.close()
        print(f"    {os.path.basename(fpath)}: {count} записів QC=1")

    recs.sort(key=lambda x: x[0])
    return recs


# ── Читання TZA ───────────────────────────────────────────────────────────────
def read_tza(tza_folder):
    """
    Завантажити рядки TZA: строки 7 і 19, TSP ≠ None.
    Колонки: День(0) | Строк(1) | TSP(2) | T(3) | WD(4) | WS(5) | ATM_код(6)
    Повертає: [(mm, day, strok, tsp, atm_code), ...]
    """
    rows = []
    folder_name = os.path.basename(tza_folder)
    for mm in range(1, 13):
        fpath = os.path.join(tza_folder, f"{mm:02d}_{folder_name}.xlsx")
        if not os.path.exists(fpath):
            print(f"    [ПРОПУСК] {os.path.basename(fpath)} — відсутній")
            continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except (BadZipFile, Exception) as e:
            print(f"    [ПРОПУСК] {os.path.basename(fpath)} — пошкоджений: {e}")
            continue

        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 7:
                continue
            try:
                day   = int(row[0])
                strok = int(float(str(row[1]).replace(',', '.')))
            except (TypeError, ValueError):
                continue
            if strok not in (7, 19) or row[2] is None:
                continue
            try:
                tsp = float(str(row[2]).replace(',', '.'))
            except (TypeError, ValueError):
                continue
            try:
                atm_code = int(row[6]) if row[6] is not None else 0
            except (TypeError, ValueError):
                atm_code = 0
            rows.append((mm, day, strok, round(tsp, 1), atm_code))
            count += 1
        wb.close()
        print(f"    {os.path.basename(fpath)}: {count} рядків")
    return rows


# ── Синхронізація ─────────────────────────────────────────────────────────────
def synchronize(cc_recs, tza_rows, window_min=30):
    """
    Зіставити кожен TZA-рядок з записами CairCloud у вікні ±window_min хв.
    Повертає список словників із усіма полями таблиці.
    """
    delta    = timedelta(minutes=window_min)
    pairs    = []
    no_match = 0
    for (mm, day, strok, tsp, atm_code) in tza_rows:
        try:
            tza_dt = datetime(2025, mm, day, strok, 0)
        except ValueError:
            continue
        window = [r for r in cc_recs
                  if tza_dt - delta <= r[0] <= tza_dt + delta]
        if not window:
            no_match += 1
            continue
        n      = len(window)
        pm10   = round(sum(r[1] for r in window) / n, 2)
        t_lst  = [r[2] for r in window if r[2] is not None]
        rh_lst = [r[3] for r in window if r[3] is not None]
        ws_lst = [r[4] for r in window if r[4] is not None]
        pairs.append({
            'date':     tza_dt.strftime('%d.%m.%Y'),
            'month_uk': MONTHS_UK[mm],
            'strok':    strok,
            'TSP':      tsp,
            'PM10':     pm10,
            'ratio':    round(pm10 / tsp, 3) if tsp > 0 else None,
            'n':        n,
            'T':        round(sum(t_lst)  / len(t_lst),  1) if t_lst  else None,
            'RH':       round(sum(rh_lst) / len(rh_lst), 1) if rh_lst else None,
            'WS':       round(sum(ws_lst) / len(ws_lst), 1) if ws_lst else None,
            'atm_code': atm_code,
            'atm_name': ATM_MAP.get(atm_code, f'код {atm_code}'),
        })
    if no_match:
        print(f"    [УВАГА] {no_match} рядків TZA без відповідності у CairCloud "
              f"(розриви у даних приладу)")
    return pairs


# ── Запис аркушу в xlsx ───────────────────────────────────────────────────────
def write_sheet(wb_out, sheet_name, title, pairs):
    ws = wb_out.create_sheet(sheet_name)

    thin       = Side(style='thin')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill   = PatternFill('solid', fgColor='D9E1F2')
    title_font = Font(name='Times New Roman', bold=True,  size=11)
    hdr_font   = Font(name='Times New Roman', bold=True,  size=10)
    data_font  = Font(name='Times New Roman',             size=10)
    center     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left       = Alignment(horizontal='left',   vertical='center')

    # Рядок 1: заголовок (об'єднані клітинки A1:L1)
    ws.merge_cells('A1:L1')
    ws['A1'] = title
    ws['A1'].font      = title_font
    ws['A1'].alignment = center

    # Рядок 2: заголовки стовпців
    headers   = ['Дата', 'Місяць', 'Строк\n(год)',
                 'TSP,\nмкг/м³', 'PM₁₀,\nмкг/м³', 'PM₁₀/TSP',
                 'n\n(зап.CC)', 'T пов.,\n°C', 'RH,\n%', 'WS,\nм/с',
                 'АТМ\nкод', 'АТМ явище']
    col_widths = [12, 11, 8, 9, 9, 8, 7, 8, 7, 7, 7, 14]

    for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(2, ci, hdr)
        c.font      = hdr_font
        c.alignment = center
        c.border    = border
        c.fill      = hdr_fill
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    # Дані (рядки з 3-го)
    for ri, p in enumerate(pairs, 3):
        vals = [p['date'], p['month_uk'], p['strok'],
                p['TSP'],  p['PM10'],    p['ratio'],
                p['n'],    p['T'],       p['RH'],    p['WS'],
                p['atm_code'], p['atm_name']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.font      = data_font
            c.border    = border
            c.alignment = left if ci == 12 else center

    ws.freeze_panes = 'A3'
    print(f"  Аркуш '{sheet_name}': {len(pairs)} рядків записано")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("=" * 60)
    print("  ЗБІРКА ДОДАТКУ 4.3")
    print(f"  Папка даних: {BASE}")
    print(f"  Вихідний файл: {OUT_4_3}")
    print("=" * 60)

    # Перевірка наявності папок
    for folder in [DIG_CC, DIG_TZA]:
        if not os.path.isdir(folder):
            print(f"\nПОМИЛКА: папка не знайдена:\n  {folder}")
            print("Перевірте що скрипт лежить у папці E:\\КЛОД\\Розіл4\\")
            sys.exit(1)

    # ── Post3 <-> pan2 ──────────────────────────────────────────
    print("\n[1/2] TZA_Post3  <->  CairCloud_pan2")
    print("  Читання CairCloud pan2 (є WS, анемометр)...")
    pan2 = read_caircloud(os.path.join(DIG_CC, 'CairCloud_DIG_pan2'), has_ws=True)
    print(f"  Всього QC=1: {len(pan2)} записів")

    print("  Читання TZA_Post3...")
    tza3 = read_tza(os.path.join(DIG_TZA, 'TZA_DIG_Post3'))
    print(f"  Всього: {len(tza3)} рядків (строки 7+19, TSP≠None)")

    print("  Синхронізація ±30 хв...")
    pairs3 = synchronize(pan2, tza3)
    print(f"  Результат: {len(pairs3)} синхронних пар\n")

    # ── Post7 <-> pan5 ──────────────────────────────────────────
    print("[2/2] TZA_Post7  <->  CairCloud_pan5")
    print("  Читання CairCloud pan5 (WS відсутній — пан5 без анемометра)...")
    pan5 = read_caircloud(os.path.join(DIG_CC, 'CairCloud_DIG_pan5'), has_ws=False)
    print(f"  Всього QC=1: {len(pan5)} записів")

    print("  Читання TZA_Post7...")
    tza7 = read_tza(os.path.join(DIG_TZA, 'TZA_DIG_Post7'))
    print(f"  Всього: {len(tza7)} рядків (строки 7+19, TSP≠None)")

    print("  Синхронізація ±30 хв...")
    pairs7 = synchronize(pan5, tza7)
    print(f"  Результат: {len(pairs7)} синхронних пар\n")

    # ── Запис файлу ─────────────────────────────────────────────
    print("Запис Додаток_4_3.xlsx...")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # видалити порожній аркуш за замовчуванням

    write_sheet(wb_out, 'Post3_pan2',
        'ДОДАТОК 4.3. Синхронізовані пари TSP (TZA_Post3) – PM₁₀ (pan2). 2025 р.',
        pairs3)
    write_sheet(wb_out, 'Post7_pan5',
        'ДОДАТОК 4.3. Синхронізовані пари TSP (TZA_Post7) – PM₁₀ (pan5). 2025 р.',
        pairs7)

    wb_out.save(OUT_4_3)

    # ── Підсумок ─────────────────────────────────────────────────
    print(f"\n{'=' * 60}")
    print(f"  ✓  ГОТОВО:  Додаток_4_3.xlsx")
    print(f"{'=' * 60}")
    print(f"\n  Post3_pan2 : {len(pairs3)} пар")
    print(f"  Post7_pan5 : {len(pairs7)} пар")
    print(f"  Всього     : {len(pairs3) + len(pairs7)} рядків\n")

    from collections import Counter
    for label, pairs in [('Post3_pan2', pairs3), ('Post7_pan5', pairs7)]:
        cnt = Counter((p['atm_code'], p['atm_name']) for p in pairs)
        print(f"  ATM коди ({label}):")
        for (code, name), n in sorted(cnt.items()):
            print(f"    {code} — {name}: {n}")
        print()
