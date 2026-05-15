# -*- coding: utf-8 -*-
"""
СКРИПТ ЗБІРКИ ТАБЛИЦІ 4.9
Зведена статистика масиву PM₁₀ по трьох сенсорах Cairnet
(pan2, pan5, pan6 — записи QC=1)
Джерело: Processed_CairCloud_DIG/
Вихід:   Таблиця_4_9.xlsx
"""
import os, sys, math
from zipfile import BadZipFile
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DIG_CC     = os.path.join(SCRIPT_DIR, "Додаток_2 Оцифровані дані", "Processed_CairCloud_DIG")
OUT_FILE   = os.path.join(SCRIPT_DIR, "Таблиця_4_9.xlsx")

SENSORS = [
    ('pan2 (пост № 2)', 'CairCloud_DIG_pan2', 1, 2, 35040, 'Включено (основна)'),
    ('pan5 (пост № 5)', 'CairCloud_DIG_pan5', 1, 2, 35040, 'Включено (верифікація)'),
    ('pan6 (пост № 6)', 'CairCloud_DIG_pan6', 1, 2, 35040, 'Включено'),
]

def load_vals(sensor_dir, ci_pm10, ci_qc):
    vals = []
    folder = os.path.basename(sensor_dir)
    for mm in range(1,13):
        fpath = os.path.join(sensor_dir, f"{mm:02d}_{folder}.xlsx")
        if not os.path.exists(fpath): continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [!] {os.path.basename(fpath)}: {e}"); continue
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if len(row) <= ci_qc or row[ci_qc] != 1: continue
            try:
                v = float(row[ci_pm10])
                if v >= 0: vals.append(v)
            except: pass
        wb.close()
    return vals

print("="*60)
print("  ЗБІРКА ТАБЛИЦІ 4.9 — Статистика PM₁₀ Cairnet")
print("="*60)

results = []
for label, subfolder, ci_pm10, ci_qc, expected, status in SENSORS:
    print(f"\n  [{label}]")
    path = os.path.join(DIG_CC, subfolder)
    vals = load_vals(path, ci_pm10, ci_qc)
    n = len(vals)
    if n == 0:
        results.append((label, 0, 0.0, '—','—','—','—', status)); continue
    mean   = round(sum(vals)/n, 1)
    sv     = sorted(vals)
    med    = round((sv[n//2-1]+sv[n//2])/2 if n%2==0 else sv[n//2], 1)
    maxv   = round(max(vals), 1)
    sigma  = round(math.sqrt(sum((x-mean)**2 for x in vals)/n), 1)
    pct    = round(100*n/expected, 1)
    print(f"    n={n}  pct={pct}%  mean={mean}  median={med}  max={maxv}  σ={sigma}")
    results.append((label, n, pct, mean, med, maxv, sigma, status))

# Стилі
thin   = Side(style='thin')
border = Border(left=thin,right=thin,top=thin,bottom=thin)
hfill  = PatternFill('solid', fgColor='D9E1F2')
tfont  = Font(name='Times New Roman', bold=True, size=11)
hfont  = Font(name='Times New Roman', bold=True, size=10)
dfont  = Font(name='Times New Roman', size=10)
center = Alignment(horizontal='center',vertical='center',wrap_text=True)
left   = Alignment(horizontal='left',vertical='center')

wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = "Таблиця 4.9"

ws.merge_cells('A1:H1')
ws['A1'] = 'Таблиця 4.9. Зведена статистика масиву PM₁₀ по трьох сенсорах Cairnet (2025 р., QC=1)'
ws['A1'].font = tfont; ws['A1'].alignment = center
ws.row_dimensions[1].height = 26

HEADERS = ['Сенсор\n(пост)','Рядків\n(QC=1)','Повнота,\n%',
           'Сер.,\nмкг/м³','Медіана,\nмкг/м³','Макс.,\nмкг/м³',
           'σ,\nмкг/м³','Статус']
WIDTHS  = [16,10,10,10,12,10,10,20]
for ci,(h,w) in enumerate(zip(HEADERS,WIDTHS),1):
    c = ws.cell(2,ci,h)
    c.font=hfont; c.alignment=center; c.border=border; c.fill=hfill
    ws.column_dimensions[get_column_letter(ci)].width=w
ws.row_dimensions[2].height=34

for ri,(label,n,pct,mean,med,maxv,sigma,status) in enumerate(results,3):
    for ci,v in enumerate([label,n,pct,mean,med,maxv,sigma,status],1):
        c=ws.cell(ri,ci,v); c.font=dfont; c.border=border
        c.alignment=left if ci in (1,8) else center

ws.merge_cells('A6:H6')
ws['A6'] = ('Примітка. Повнота розрахована відносно теоретичного максимуму '
            '35 040 записів (365 діб × 24 год × 4 записи/год). '
            'QC=1 — прийнятий запис. Джерело: Processed_CairCloud_DIG/ (build_Таблиця_4_9.py).')
ws['A6'].font = Font(name='Times New Roman',italic=True,size=9)
ws['A6'].alignment = Alignment(wrap_text=True,vertical='top')
ws.row_dimensions[6].height=45
ws.freeze_panes='A3'
wb_out.save(OUT_FILE)
print(f"\n  ✓ ТАБЛИЦЯ 4.9: {OUT_FILE}")
