# -*- coding: utf-8 -*-
"""
СКРИПТ ЗБІРКИ ТАБЛИЦІ 4.10
Сезонний аналіз масиву синхронних пар
(TSP — TZA_Post3 / PM₁₀ — Cairnet pan2)
Джерело: Додаток_4_3.xlsx (аркуш Post3_pan2)
Вихід:   Таблиця_4_10.xlsx
"""
import os, sys
from zipfile import BadZipFile
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_FILE   = os.path.join(SCRIPT_DIR, "Додаток_4_3.xlsx")
OUT_FILE   = os.path.join(SCRIPT_DIR, "Таблиця_4_10.xlsx")

# Колонки (0-based): Date(0) Month(1) Hour(2) TSP(3) PM10(4)
#                    PM10/TSP(5) n(6) T(7) RH(8) WS(9) ATM_code(10) ATM(11)
COL_TSP, COL_PM10, COL_T, COL_RH = 3, 4, 7, 8

# Сезони: (назва, місяці_UK, місяці_числа)
SEASONS = [
    ('Зима',  'Грд–Лют',  {12, 1, 2}),
    ('Весна', 'Бер–Тра',  {3, 4, 5}),
    ('Літо',  'Чер–Сер',  {6, 7, 8}),
    ('Осінь', 'Вер–Лис',  {9, 10, 11}),
]

MONTH_NUM = {
    'Січень':1,'Лютий':2,'Березень':3,'Квітень':4,'Травень':5,'Червень':6,
    'Липень':7,'Серпень':8,'Вересень':9,'Жовтень':10,'Листопад':11,'Грудень':12
}

print("="*60)
print("  ЗБІРКА ТАБЛИЦІ 4.10 — Сезонний аналіз")
print("="*60)

wb_src = openpyxl.load_workbook(SRC_FILE, read_only=True, data_only=True)
ws_src = wb_src['Post3_pan2']

season_data = {s[0]: {'tsp':[],'pm10':[],'rh':[],'t':[]} for s in SEASONS}

for row in ws_src.iter_rows(min_row=3, values_only=True):
    if not row or row[0] is None: continue
    month_str = row[1]
    if month_str not in MONTH_NUM: continue
    mm = MONTH_NUM[month_str]
    tsp  = row[COL_TSP]
    pm10 = row[COL_PM10]
    t    = row[COL_T]
    rh   = row[COL_RH]
    for sname, slabel, smonths in SEASONS:
        if mm in smonths:
            if tsp  is not None: season_data[sname]['tsp'].append(float(tsp))
            if pm10 is not None: season_data[sname]['pm10'].append(float(pm10))
            if t    is not None: season_data[sname]['t'].append(float(t))
            if rh   is not None: season_data[sname]['rh'].append(float(rh))
wb_src.close()

results = []
for sname, slabel, _ in SEASONS:
    d = season_data[sname]
    n = len(d['tsp'])
    mean_tsp  = round(sum(d['tsp'])/len(d['tsp']),1)   if d['tsp']  else '—'
    mean_pm10 = round(sum(d['pm10'])/len(d['pm10']),1) if d['pm10'] else '—'
    mean_rh   = round(sum(d['rh'])/len(d['rh']),0)     if d['rh']   else '—'
    mean_t    = round(sum(d['t'])/len(d['t']),1)        if d['t']    else '—'
    results.append((sname, slabel, n, mean_tsp, mean_pm10,
                    int(mean_rh) if mean_rh!='—' else '—', mean_t))
    print(f"  {sname:6}: n={n:3}  TSP={mean_tsp}  PM10={mean_pm10}  RH={mean_rh}%  T={mean_t}°C")

# Стилі
thin   = Side(style='thin')
border = Border(left=thin,right=thin,top=thin,bottom=thin)
hfill  = PatternFill('solid', fgColor='D9E1F2')
SFILLS = ['DCE6F1','E2EFDA','FFF2CC','FCE4D6']
tfont  = Font(name='Times New Roman', bold=True, size=11)
hfont  = Font(name='Times New Roman', bold=True, size=10)
dfont  = Font(name='Times New Roman', size=10)
center = Alignment(horizontal='center',vertical='center',wrap_text=True)
left   = Alignment(horizontal='left',vertical='center')

wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = "Таблиця 4.10"

ws.merge_cells('A1:G1')
ws['A1'] = ('Таблиця 4.10. Сезонний аналіз масиву синхронних пар '
            'TSP (TZA_Post3) – PM₁₀ (pan2), 2025 р.')
ws['A1'].font = tfont; ws['A1'].alignment = center
ws.row_dimensions[1].height = 26

HEADERS = ['Сезон','Місяці','Пар\n(кількість)',
           'Сер. TSP,\nмкг/м³','Сер. PM₁₀,\nмкг/м³',
           'Сер. RH,\n%','Сер. T,\n°C']
WIDTHS  = [10,10,12,14,14,10,10]
for ci,(h,w) in enumerate(zip(HEADERS,WIDTHS),1):
    c = ws.cell(2,ci,h)
    c.font=hfont; c.alignment=center; c.border=border; c.fill=hfill
    ws.column_dimensions[get_column_letter(ci)].width=w
ws.row_dimensions[2].height=34

for ri,(sname,slabel,n,tsp,pm10,rh,t) in enumerate(results,3):
    sfill = PatternFill('solid', fgColor=SFILLS[ri-3])
    for ci,v in enumerate([sname,slabel,n,tsp,pm10,rh,t],1):
        c=ws.cell(ri,ci,v); c.font=dfont; c.border=border
        c.alignment=left if ci<=2 else center
        c.fill=sfill

ws.merge_cells('A7:G7')
ws['A7'] = ('Примітка. Дані з Додаток_4_3.xlsx (аркуш Post3_pan2). '
            'Зима: грудень–лютий; Весна: березень–травень; '
            'Літо: червень–серпень; Осінь: вересень–листопад. '
            'Джерело: build_Таблиця_4_10.py.')
ws['A7'].font = Font(name='Times New Roman',italic=True,size=9)
ws['A7'].alignment = Alignment(wrap_text=True,vertical='top')
ws.row_dimensions[7].height=45
ws.freeze_panes='A3'
wb_out.save(OUT_FILE)
print(f"\n  ✓ ТАБЛИЦЯ 4.10: {OUT_FILE}")
