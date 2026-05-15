# -*- coding: utf-8 -*-
"""
===========================================================
  СКРИПТ ЗБІРКИ ТАБЛИЦІ 4.6
  Повнота ряду даних сенсора CairCloud pan5 по місяцях

  Таблиця 4.6 — Розділ_4, підрозділ 4.4.1
  Джерело даних: Processed_CairCloud_DIG/CairCloud_DIG_pan5/

  Вихідний файл: Таблиця_4_6.xlsx  (у тій самій папці)

  ЗАПУСК:
      python build_Таблиця_4_6.py

  ЗАЛЕЖНОСТІ: openpyxl  (pip install openpyxl)
===========================================================
"""

import os, sys, calendar
from zipfile import BadZipFile

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ПОМИЛКА: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PAN5_DIR   = os.path.join(SCRIPT_DIR,
             "Додаток_2 Оцифровані дані",
             "Processed_CairCloud_DIG",
             "CairCloud_DIG_pan5")
OUT_FILE   = os.path.join(SCRIPT_DIR, "Таблиця_4_6.xlsx")
YEAR       = 2025

MONTHS_UK  = ['', 'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 'Червень',
              'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень']

# Індекси колонок у pan5 DIG (0-based):
# Timestamp(0) | PM10(1) | QC(2) | T(3) | ...
CC_PM10 = 1
CC_QC   = 2

print("=" * 60)
print("  ЗБІРКА ТАБЛИЦІ 4.6 — Повнота ряду pan5")
print(f"  Дані: {PAN5_DIR}")
print("=" * 60)

rows_data = []
total_qc1 = 0
total_all  = 0
total_exp  = 0

for mm in range(1, 13):
    days_in_month = calendar.monthrange(YEAR, mm)[1]
    # CairCloud записує кожні 15 хв → 4 записи/год → 96/добу
    expected = days_in_month * 24 * 4

    fpath = os.path.join(PAN5_DIR, f"{mm:02d}_CairCloud_DIG_pan5.xlsx")
    if not os.path.exists(fpath):
        print(f"  [ПРОПУСК] {os.path.basename(fpath)} — відсутній")
        rows_data.append((MONTHS_UK[mm], days_in_month, 0, 0, expected, 0.0, 0.0))
        total_exp += expected
        continue

    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except (BadZipFile, Exception) as e:
        print(f"  [ПРОПУСК] {os.path.basename(fpath)} — пошкоджений: {e}")
        rows_data.append((MONTHS_UK[mm], days_in_month, 0, 0, expected, 0.0, 0.0))
        total_exp += expected
        continue

    ws = wb.active
    n_all = 0
    n_qc1 = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= CC_QC:
            continue
        n_all += 1
        if row[CC_QC] == 1:
            n_qc1 += 1
    wb.close()

    pct_qc1_all = round(100 * n_qc1 / n_all, 1)  if n_all  else 0.0
    pct_qc1_exp = round(100 * n_qc1 / expected, 1) if expected else 0.0
    rows_data.append((MONTHS_UK[mm], days_in_month, n_all, n_qc1, expected,
                      pct_qc1_all, pct_qc1_exp))
    total_qc1 += n_qc1
    total_all  += n_all
    total_exp  += expected
    print(f"  {MONTHS_UK[mm]:10}: QC=1={n_qc1:5}  всього={n_all:5}  {pct_qc1_all:5.1f}%")

pct_tot = round(100 * total_qc1 / total_all,  1) if total_all  else 0.0
pct_exp = round(100 * total_qc1 / total_exp,  1) if total_exp  else 0.0

# ── Запис xlsx ────────────────────────────────────────────────────────────────
thin   = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill  = PatternFill('solid', fgColor='D9E1F2')
tfont  = Font(name='Times New Roman', bold=True,  size=11)
hfont  = Font(name='Times New Roman', bold=True,  size=10)
dfont  = Font(name='Times New Roman',             size=10)
bfont  = Font(name='Times New Roman', bold=True,  size=10)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left',   vertical='center')

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Таблиця 4.6"

# Рядок 1: заголовок
ws_out.merge_cells('A1:G1')
ws_out['A1'] = ('Таблиця 4.6. Повнота ряду даних сенсора CairCloud pan5 '
                'по місяцях (2025 р., Кривий Ріг)')
ws_out['A1'].font      = tfont
ws_out['A1'].alignment = center
ws_out.row_dimensions[1].height = 26

# Рядок 2: заголовки стовпців
HEADERS = [
    'Місяць',
    'Діб у\nмісяці',
    'Всього\nзаписів',
    'Записів\nQC=1',
    'Очікувано\n(15-хв. крок)',
    '% QC=1\nвід всього',
    '% QC=1\nвід очікуваного',
]
WIDTHS = [12, 8, 10, 10, 14, 10, 14]

for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
    c = ws_out.cell(2, ci, h)
    c.font = hfont; c.alignment = center; c.border = border; c.fill = hfill
    ws_out.column_dimensions[get_column_letter(ci)].width = w
ws_out.row_dimensions[2].height = 34

# Рядки 3-14: дані
for ri, (month, days, n_all, n_qc1, exp, pct_all, pct_exp) in enumerate(rows_data, 3):
    vals = [month, days, n_all, n_qc1, exp, pct_all, pct_exp]
    for ci, v in enumerate(vals, 1):
        c = ws_out.cell(ri, ci, v)
        c.font = dfont; c.border = border
        c.alignment = left if ci == 1 else center

# Рядок 15: підсумок
ws_out.row_dimensions[15].height = 18
totals = ['РАЗОМ', '', total_all, total_qc1, total_exp, pct_tot, pct_exp]
for ci, v in enumerate(totals, 1):
    c = ws_out.cell(15, ci, v)
    c.font = bfont; c.border = border
    c.alignment = left if ci == 1 else center
    c.fill = PatternFill('solid', fgColor='F2F2F2')

# Рядок 16: примітка
ws_out.merge_cells('A16:G16')
ws_out['A16'] = ('Примітка. QC=1 — прийнятий запис (прапор контролю якості). '
                 'Очікувана кількість розрахована як кількість діб × 24 год × 4 записи/год '
                 '(крок 15 хвилин). Джерело даних: Processed_CairCloud_DIG/CairCloud_DIG_pan5/')
ws_out['A16'].font = Font(name='Times New Roman', italic=True, size=9)
ws_out['A16'].alignment = Alignment(wrap_text=True, vertical='top')
ws_out.row_dimensions[16].height = 40

ws_out.freeze_panes = 'A3'
wb_out.save(OUT_FILE)

print(f"\n{'=' * 60}")
print(f"  ✓  ТАБЛИЦЯ 4.6 ЗБЕРЕЖЕНА: {OUT_FILE}")
print(f"{'=' * 60}")
print(f"  Всього QC=1 : {total_qc1}  ({pct_tot}% від всіх записів)")
print(f"  Очікувано   : {total_exp}  QC=1/очік = {pct_exp}%")
