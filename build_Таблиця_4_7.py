# -*- coding: utf-8 -*-
"""
===========================================================
  СКРИПТ ЗБІРКИ ТАБЛИЦІ 4.7
  Попарні відношення місячних середніх PM₁₀
  між сенсорами pan2, pan5, pan6 (2025 р.)

  Таблиця 4.7 — Розділ_4, підрозділ 4.4.2
  Джерело: Processed_CairCloud_DIG/ (pan2, pan5, pan6)

  Вихідний файл: Таблиця_4_7.xlsx  (у тій самій папці)

  ЗАПУСК:
      python build_Таблиця_4_7.py

  ЗАЛЕЖНОСТІ: openpyxl  (pip install openpyxl)
===========================================================
"""

import os, sys
from zipfile import BadZipFile
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ПОМИЛКА: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DIG_CC     = os.path.join(SCRIPT_DIR,
             "Додаток_2 Оцифровані дані",
             "Processed_CairCloud_DIG")
OUT_FILE   = os.path.join(SCRIPT_DIR, "Таблиця_4_7.xlsx")

MONTHS_UK = ['', 'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 'Червень',
             'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень']

# Конфігурація датчиків: (ім'я, підпапка, індекс_PM10, індекс_QC)
SENSORS = [
    ('pan2', 'CairCloud_DIG_pan2', 1, 2),
    ('pan5', 'CairCloud_DIG_pan5', 1, 2),
    ('pan6', 'CairCloud_DIG_pan6', 1, 2),
]


def load_monthly_means(sensor_dir, cc_pm10, cc_qc):
    """Обчислити середнє PM₁₀ (QC=1) по місяцях. Повертає {mm: mean_or_None}."""
    means = {}
    folder_name = os.path.basename(sensor_dir)
    for mm in range(1, 13):
        fpath = os.path.join(sensor_dir, f"{mm:02d}_{folder_name}.xlsx")
        if not os.path.exists(fpath):
            means[mm] = None
            continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except (BadZipFile, Exception) as e:
            print(f"  [!] {os.path.basename(fpath)}: {e}")
            means[mm] = None
            continue
        ws = wb.active
        vals = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= cc_qc or row[cc_qc] != 1:
                continue
            try:
                v = float(row[cc_pm10])
                if v >= 0:
                    vals.append(v)
            except (TypeError, ValueError):
                pass
        wb.close()
        means[mm] = round(sum(vals) / len(vals), 2) if vals else None
        print(f"  {folder_name} {MONTHS_UK[mm]:10}: "
              f"n={len(vals):5}  mean={means[mm]:.2f}" if means[mm] else
              f"  {folder_name} {MONTHS_UK[mm]:10}: —")
    return means


print("=" * 60)
print("  ЗБІРКА ТАБЛИЦІ 4.7 — Місячні середні PM₁₀ і відношення")
print("=" * 60)

all_means = {}
for name, subfolder, ci_pm10, ci_qc in SENSORS:
    print(f"\n  [{name}]")
    sensor_path = os.path.join(DIG_CC, subfolder)
    all_means[name] = load_monthly_means(sensor_path, ci_pm10, ci_qc)

# Річні середні
annual = {}
for name, subfolder, ci_pm10, ci_qc in SENSORS:
    sensor_path = os.path.join(DIG_CC, subfolder)
    folder_name = subfolder
    vals_year = []
    for mm in range(1, 13):
        fpath = os.path.join(sensor_path, f"{mm:02d}_{folder_name}.xlsx")
        if not os.path.exists(fpath):
            continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception:
            continue
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= ci_qc or row[ci_qc] != 1:
                continue
            try:
                v = float(row[ci_pm10])
                if v >= 0:
                    vals_year.append(v)
            except (TypeError, ValueError):
                pass
        wb.close()
    annual[name] = round(sum(vals_year) / len(vals_year), 2) if vals_year else None

# ── Запис xlsx ────────────────────────────────────────────────────────────────
thin   = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill  = PatternFill('solid', fgColor='D9E1F2')
yfill  = PatternFill('solid', fgColor='FFF2CC')   # підсвітка відношень
tfont  = Font(name='Times New Roman', bold=True, size=11)
hfont  = Font(name='Times New Roman', bold=True, size=10)
dfont  = Font(name='Times New Roman',            size=10)
bfont  = Font(name='Times New Roman', bold=True, size=10)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left',   vertical='center')

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Таблиця 4.7"

# Рядок 1: заголовок
ws_out.merge_cells('A1:H1')
ws_out['A1'] = ('Таблиця 4.7. Місячні середні значення PM₁₀ (мкг/м³) '
                'та попарні відношення між сенсорами (2025 р., Кривий Ріг)')
ws_out['A1'].font      = tfont
ws_out['A1'].alignment = center
ws_out.row_dimensions[1].height = 26

# Рядок 2: заголовки
HEADERS = [
    'Місяць',
    'pan2,\nмкг/м³',
    'pan5,\nмкг/м³',
    'pan6,\nмкг/м³',
    'pan6/pan2',
    'pan2/pan5',
    'Δ pan6–pan2,\nмкг/м³',
    'Δ pan2–pan5,\nмкг/м³',
]
WIDTHS = [12, 10, 10, 10, 10, 10, 13, 13]

for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
    c = ws_out.cell(2, ci, h)
    c.font = hfont; c.alignment = center; c.border = border; c.fill = hfill
    ws_out.column_dimensions[get_column_letter(ci)].width = w
ws_out.row_dimensions[2].height = 34

# Рядки 3-14: дані по місяцях
for ri, mm in enumerate(range(1, 13), 3):
    m2 = all_means['pan2'][mm]
    m5 = all_means['pan5'][mm]
    m6 = all_means['pan6'][mm]

    r62  = round(m6 / m2, 3) if (m2 and m6) else None
    r25  = round(m2 / m5, 3) if (m2 and m5) else None
    d62  = round(m6 - m2, 2) if (m2 and m6) else None
    d25  = round(m2 - m5, 2) if (m2 and m5) else None

    vals = [MONTHS_UK[mm],
            m2 if m2 else '—',
            m5 if m5 else '—',
            m6 if m6 else '—',
            r62 if r62 else '—',
            r25 if r25 else '—',
            d62 if d62 else '—',
            d25 if d25 else '—']

    for ci, v in enumerate(vals, 1):
        c = ws_out.cell(ri, ci, v)
        c.font = dfont; c.border = border
        c.alignment = left if ci == 1 else center
        # Підсвітка відношень > 1.5 (значна різниця)
        if ci in (5, 6) and isinstance(v, float) and v > 1.5:
            c.fill = yfill

# Рядок 15: річні середні
ws_out.row_dimensions[15].height = 18
a2 = annual['pan2']; a5 = annual['pan5']; a6 = annual['pan6']
ra62 = round(a6/a2, 3) if (a2 and a6) else None
ra25 = round(a2/a5, 3) if (a2 and a5) else None
da62 = round(a6-a2, 2) if (a2 and a6) else None
da25 = round(a2-a5, 2) if (a2 and a5) else None

annual_row = ['Річне середнє',
              a2 if a2 else '—',
              a5 if a5 else '—',
              a6 if a6 else '—',
              ra62 if ra62 else '—',
              ra25 if ra25 else '—',
              da62 if da62 else '—',
              da25 if da25 else '—']

for ci, v in enumerate(annual_row, 1):
    c = ws_out.cell(15, ci, v)
    c.font = bfont; c.border = border
    c.alignment = left if ci == 1 else center
    c.fill = PatternFill('solid', fgColor='F2F2F2')

# Рядок 16: примітка
ws_out.merge_cells('A16:H16')
ws_out['A16'] = ('Примітка. Середні розраховані за записами з QC=1. '
                 'pan6/pan2 — відношення середніх між постами № 6 та № 2; '
                 'pan2/pan5 — між постами № 2 та № 5. '
                 'Жовтим виділені місяці з відношенням > 1,5. '
                 'Джерело: Processed_CairCloud_DIG/ (build_Таблиця_4_7.py).')
ws_out['A16'].font = Font(name='Times New Roman', italic=True, size=9)
ws_out['A16'].alignment = Alignment(wrap_text=True, vertical='top')
ws_out.row_dimensions[16].height = 40

ws_out.freeze_panes = 'A3'
wb_out.save(OUT_FILE)

print(f"\n{'=' * 60}")
print(f"  ✓  ТАБЛИЦЯ 4.7 ЗБЕРЕЖЕНА: {OUT_FILE}")
print(f"{'=' * 60}")
print(f"  Річні середні: pan2={a2}  pan5={a5}  pan6={a6} мкг/м³")
print(f"  Відношення річних: pan6/pan2={ra62}  pan2/pan5={ra25}")
