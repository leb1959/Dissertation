# -*- coding: utf-8 -*-
"""
СКРИПТ ЗБІРКИ ТАБЛИЦІ 4.13
Частота потенційних викидів за методом Тюкі (місячна статистика)

Методологія:
  TZA_Post3: строки 7 та 19 (VALID_STROK={7,19}), TSP>0 (нульові = пропущені виміри).
              Межі Тюкі розраховуються ЛОКАЛЬНО по кожному місяцю (k=1.5).
  Cairnet pan2: QC=1, PM10≥0. Межі Тюкі ЛОКАЛЬНО по місяцю (k=1.5).
  Метеообумовлені,%: ручна оцінка на основі журналів ТЗА
                     (позначки «туман», «хмарність», «опади»).
  Залишено у вибірці: всі місяці залишені (викиди є реальними подіями).

Вхідні дані:
  Processed_TZA_DIG/TZA_DIG_Post3/MM_TZA_DIG_Post3.xlsx
  Processed_CairCloud_DIG/CairCloud_DIG_pan2/MM_CairCloud_DIG_pan2.xlsx

Вихід: Таблиця_4_13.xlsx
"""
import os, sys
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TZA3_DIR   = os.path.join(SCRIPT_DIR, "Додаток_2 Оцифровані дані",
                           "Processed_TZA_DIG", "TZA_DIG_Post3")
PAN2_DIR   = os.path.join(SCRIPT_DIR, "Додаток_2 Оцифровані дані",
                           "Processed_CairCloud_DIG", "CairCloud_DIG_pan2")
OUT_FILE   = os.path.join(SCRIPT_DIR, "Таблиця_4_13.xlsx")

MONTHS_UK  = ['','Січень','Лютий','Березень','Квітень','Травень','Червень',
              'Липень','Серпень','Вересень','Жовтень','Листопад','Грудень']
VALID_STROK = {7, 19}

# Метеообумовлені,% — ручна оцінка з журналів ТЗА
# (частка "викидів" TZA що збігаються з метеопозначками у журналах)
METEO_PCT = {1:100, 2:89, 3:100, 4:100, 5:100, 6:100,
             7:100, 8:100, 9:100, 10:100, 11:100, 12:91}

def quartile_inc(data, p):
    """QUARTILE.INC (Excel-сумісний)"""
    s = sorted(data); n = len(s)
    pos = p*(n-1); lo = int(pos); hi = lo+1
    if hi >= n: return s[-1]
    return s[lo] + (pos-lo)*(s[hi]-s[lo])

def tukey_local(vals, k=1.5):
    """Локальний Тюкі; повертає (%, n_out, n_total)"""
    if len(vals) < 4: return None, 0, len(vals)
    q1 = quartile_inc(vals, 0.25)
    q3 = quartile_inc(vals, 0.75)
    iqr = q3 - q1
    lo  = q1 - k*iqr
    hi  = q3 + k*iqr
    n_out = sum(1 for v in vals if v < lo or v > hi)
    pct = round(n_out / len(vals) * 100, 1) if vals else None
    return pct, n_out, len(vals)

print("="*60)
print("  ЗБІРКА ТАБЛИЦІ 4.13 — Аналіз викидів (Тюкі)")
print("  TZA: локальний Тюкі на TSP>0 | pan2: QC=1, PM10≥0")
print("="*60)

rows_data  = []
total_tza_out = 0; total_tza_n = 0
total_pm_out  = 0; total_pm_n  = 0

for mm in range(1, 13):
    # --- TZA Post3 ---
    tza_path = os.path.join(TZA3_DIR, f"{mm:02d}_TZA_DIG_Post3.xlsx")
    tza_vals = []
    if os.path.exists(tza_path):
        wb = openpyxl.load_workbook(tza_path, read_only=True, data_only=True)
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if row and row[1] in VALID_STROK and row[2] is not None:
                try:
                    v = float(row[2])
                    if v > 0: tza_vals.append(v)
                except: pass
        wb.close()
    else:
        print(f"  [!] TZA файл не знайдено: {tza_path}")

    # --- CairCloud pan2 ---
    pan2_path = os.path.join(PAN2_DIR, f"{mm:02d}_CairCloud_DIG_pan2.xlsx")
    pm10_vals = []
    if os.path.exists(pan2_path):
        wb = openpyxl.load_workbook(pan2_path, read_only=True, data_only=True)
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if row and row[2] == 1 and row[1] is not None:
                try:
                    v = float(row[1])
                    if v >= 0: pm10_vals.append(v)
                except: pass
        wb.close()
    else:
        print(f"  [!] pan2 файл не знайдено: {pan2_path}")

    pct_tza, out_tza, n_tza = tukey_local(tza_vals)
    pct_pm,  out_pm,  n_pm  = tukey_local(pm10_vals)

    # Для жовтня-листопада PM10 менше 400 рядків — ненадійно, ставимо "—"
    pm_display = f"{pct_pm}" if n_pm >= 400 else None

    total_tza_out += out_tza; total_tza_n += n_tza
    if n_pm >= 400: total_pm_out += out_pm; total_pm_n += n_pm

    meteo = METEO_PCT.get(mm, 100)

    rows_data.append((MONTHS_UK[mm], pct_tza, pm_display, meteo, 'Так',
                      n_tza, out_tza, n_pm, out_pm))
    print(f"  {MONTHS_UK[mm]:10}: TZA={pct_tza}% ({out_tza}/{n_tza})  "
          f"PM10={pm_display}% ({out_pm}/{n_pm})  Метео={meteo}%")

pct_total_tza = round(total_tza_out/total_tza_n*100,1) if total_tza_n else 0
pct_total_pm  = round(total_pm_out/total_pm_n*100,1)   if total_pm_n  else 0

print(f"\n  Разом TZA: {total_tza_out} з {total_tza_n} = {pct_total_tza}%")
print(f"  Разом PM₁₀ (міс. ≥400 рядків): {total_pm_out} з {total_pm_n} = {pct_total_pm}%")

# ── Запис xlsx ─────────────────────────────────────────────
thin   = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill  = PatternFill('solid', fgColor='D9E1F2')
tfont  = Font(name='Times New Roman', bold=True,   size=11)
hfont  = Font(name='Times New Roman', bold=True,   size=10)
dfont  = Font(name='Times New Roman',              size=10)
bfont  = Font(name='Times New Roman', bold=True,   size=10)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left',   vertical='center')

wb_out = openpyxl.Workbook()
ws     = wb_out.active
ws.title = "Таблиця 4.13"

ws.merge_cells('A1:E1')
ws['A1'] = ('Таблиця 4.13. Частота потенційних викидів за методом Тюкі '
            '(місячна статистика, TZA_Post3 / Cairnet pan2). '
            'Локальні межі k=1.5; TZA — лише TSP>0; PM₁₀ — QC=1.')
ws['A1'].font      = tfont
ws['A1'].alignment = center
ws.row_dimensions[1].height = 36

HEADERS = ['Місяць','Викидів\nТЗА,%','Викидів\nPM₁₀,%','Метео-\nобумовлені,%','Залишено\nу вибірці']
WIDTHS  = [12, 11, 11, 14, 13]
for ci,(h,w) in enumerate(zip(HEADERS,WIDTHS),1):
    c = ws.cell(2, ci, h)
    c.font=hfont; c.alignment=center; c.border=border; c.fill=hfill
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[2].height = 36

for ri,(month, pt, pp, meteo, zal, *_) in enumerate(rows_data, 3):
    vals = [month, f'{pt}' if pt is not None else '—',
            f'{pp}' if pp is not None else '—', str(meteo), zal]
    for ci,v in enumerate(vals, 1):
        c = ws.cell(ri, ci, v)
        c.font=dfont; c.border=border
        c.alignment = left if ci==1 else center

ws.merge_cells('A16:E16')
ws['A16'] = (f'Примітка. TZA: метод Тюкі (k=1.5) на ненульових значеннях TSP '
             f'(строки 7 та 19); всього {total_tza_n} вимірювань, '
             f'{total_tza_out} потенційних викидів ({pct_total_tza}%). '
             f'PM₁₀: метод Тюкі (k=1.5) на QC=1 записах Cairnet pan2; '
             f'місяці з <400 рядків (жовт., лист.) позначені «—». '
             f'Метеообумовлені,% — оцінка за журналами ТЗА. '
             f'Джерело: build_Таблиця_4_13.py.')
ws['A16'].font      = Font(name='Times New Roman', italic=True, size=9)
ws['A16'].alignment = Alignment(wrap_text=True, vertical='top')
ws.row_dimensions[16].height = 50

wb_out.save(OUT_FILE)
print(f'\n  Таблиця_4_13.xlsx збережено: {OUT_FILE}')
