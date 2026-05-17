"""
build_Таблиця_5_2_1.py
Генерує Таблиця_5_2_1.xlsx — міжпостові коефіцієнти кореляції PM10 за місяцями
(POST2, POST6, pan5) — дані Розділу 5 дисертації
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "Таблиця_5_2_1.xlsx"

def hdr(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=True, color="FFFFFF", name='Times New Roman', size=11)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def dat(ws, r, c, v, fill=None, bold=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, name='Times New Roman', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Таблиця 5.2.1"

ws.merge_cells('A1:E1')
ws['A1'].value = "Таблиця 5.2.1 – Узагальнення міжпостових коефіцієнтів кореляції PM₁₀ за місяцями"
ws['A1'].font = Font(bold=True, name='Times New Roman', size=12)
ws['A1'].alignment = Alignment(horizontal='center', wrap_text=True)
ws.row_dimensions[1].height = 30

for c, h in enumerate(['Місяць', 'r (POST2–POST6)', 'r (POST2–pan5)', 'r (POST6–pan5)', 'Примітка'], 1):
    hdr(ws, 2, c, h)
ws.row_dimensions[2].height = 35

DATA = [
    ('Січень',   0.66, 0.58, 0.59, 'помірно висока узгодженість'),
    ('Лютий',    0.74, 0.76, 0.80, 'висока узгодженість'),
    ('Березень', 0.62, 0.62, 0.79, 'часткове покриття pan5'),
    ('Квітень',  0.77, 0.54, 0.66, 'обмежена репрезентативність'),
    ('Травень',  0.69, 0.40, 0.49, 'послаблення узгодженості'),
    ('Червень',  0.83, 0.76, 0.78, 'висока узгодженість'),
    ('Липень',   0.77, 0.64, 0.56, 'локальні літні відмінності'),
    ('Серпень',  0.66, 0.53, 0.73, 'часткове відновлення узгодженості'),
    ('Вересень', 0.53, 0.64, 0.70, 'помірна узгодженість'),
    ('Жовтень',  0.82, 0.67, 0.76, 'часткове покриття'),
    ('Листопад', '—', '—', '—',   'даних недостатньо'),
    ('Грудень',  0.93, 0.88, 0.90, 'найкраща узгодженість'),
]

for r, (month, r1, r2, r3, note) in enumerate(DATA, 3):
    fill = "C6EFCE" if isinstance(r1, float) and r1 >= 0.80 else ("FFEB9C" if isinstance(r1, float) and r1 < 0.60 else None)
    dat(ws, r, 1, month, bold=True)
    dat(ws, r, 2, r1, fill=fill)
    dat(ws, r, 3, r2)
    dat(ws, r, 4, r3)
    dat(ws, r, 5, note)

for i, w in enumerate([15, 14, 14, 14, 35], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print(f"Збережено: {OUT}")
