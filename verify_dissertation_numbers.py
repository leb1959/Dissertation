# -*- coding: utf-8 -*-
"""
===========================================================
  ВЕРИФІКАЦІЯ ЧИСЛОВИХ ДАНИХ ДИСЕРТАЦІЇ (Розділ 4)
  Перевірка всіх ключових цифр з реальних DIG-файлів

  Питання, що перевіряються:
    Q1 — Середня швидкість вітру 2.5 м/с (pan2)
    Q3 — Статистика по всіх датчиках (підрозділ 4.4.2)
    Q4 — QC TZA_Post7: 621 / 48 / 573
    Q5 — Таблиця 4.6: повнота ряду pan5 по місяцях
    Q6 — max PM₁₀ pan5=45, pan6 статистика
    Q7 — Таблиця 4.7: коефіцієнти pan6/pan2, pan2/pan5 по місяцях

  ЗАПУСК: python verify_dissertation_numbers.py
===========================================================
"""

import os
import sys
from datetime import datetime
from zipfile import BadZipFile
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ПОМИЛКА: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE    = os.path.join(SCRIPT_DIR, "Додаток_2 Оцифровані дані")
DIG_CC  = os.path.join(BASE, "Processed_CairCloud_DIG")
DIG_TZA = os.path.join(BASE, "Processed_TZA_DIG")

MONTHS_UK = ['', 'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 'Червень',
             'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень']

SEP = "=" * 65

def load_caircloud(pan_folder, has_ws=True):
    """
    Повертає: {mm: [{'dt', 'pm10', 't', 'rh', 'ws', 'qc'}, ...]}
    Завантажує ВСІ записи (не тільки QC=1) для підрахунку повноти.
    """
    CC_PM10, CC_QC, CC_T, CC_RH = 1, 2, 3, 5
    CC_WS = 9 if has_ws else None
    data = defaultdict(list)
    folder_name = os.path.basename(pan_folder)
    for mm in range(1, 13):
        fpath = os.path.join(pan_folder, f"{mm:02d}_{folder_name}.xlsx")
        if not os.path.exists(fpath):
            continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except (BadZipFile, Exception):
            continue
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= CC_QC:
                continue
            try:
                pm10 = float(row[CC_PM10]) if row[CC_PM10] is not None else None
            except (TypeError, ValueError):
                pm10 = None
            qc = row[CC_QC]
            ts = row[0]
            if isinstance(ts, datetime):
                dt = ts
            elif isinstance(ts, str):
                dt = None
                for fmt in ['%Y-%m-%d %H:%M', '%d.%m.%Y %H:%M', '%Y-%m-%d %H:%M:%S']:
                    try:
                        dt = datetime.strptime(ts.strip(), fmt); break
                    except ValueError:
                        pass
                if dt is None:
                    continue
            else:
                continue
            t_val = rh_val = ws_val = None
            try: t_val  = float(row[CC_T])  if len(row) > CC_T  else None
            except: pass
            try: rh_val = float(row[CC_RH]) if len(row) > CC_RH else None
            except: pass
            if CC_WS and len(row) > CC_WS:
                try: ws_val = float(row[CC_WS])
                except: pass
            data[mm].append({'dt': dt, 'pm10': pm10, 't': t_val,
                             'rh': rh_val, 'ws': ws_val, 'qc': qc})
        wb.close()
    return data


def load_tza_full(tza_folder):
    """
    Повертає всі рядки TZA (НЕ фільтруючи строки/TSP) для підрахунку QC.
    [(mm, day, strok, tsp, atm_code, row_raw), ...]
    """
    rows_all = []
    folder_name = os.path.basename(tza_folder)
    for mm in range(1, 13):
        fpath = os.path.join(tza_folder, f"{mm:02d}_{folder_name}.xlsx")
        if not os.path.exists(fpath):
            continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except (BadZipFile, Exception):
            continue
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 3:
                continue
            rows_all.append((mm, row))
        wb.close()
    return rows_all


print(SEP)
print("  ВЕРИФІКАЦІЯ ЧИСЛОВИХ ДАНИХ ДИСЕРТАЦІЇ")
print(SEP)

# ══════════════════════════════════════════════════════════════
# Q1 — СЕРЕДНЯ ШВИДКІСТЬ ВІТРУ (pan2)
# ══════════════════════════════════════════════════════════════
print("\n" + SEP)
print("  Q1 — СЕРЕДНЯ ШВИДКІСТЬ ВІТРУ WS (pan2, тільки QC=1)")
print(SEP)

pan2_data = load_caircloud(os.path.join(DIG_CC, 'CairCloud_DIG_pan2'), has_ws=True)
ws_all = []
ws_by_month = {}
for mm in range(1, 13):
    ws_m = [r['ws'] for r in pan2_data[mm] if r['qc'] == 1 and r['ws'] is not None]
    ws_by_month[mm] = ws_m
    ws_all.extend(ws_m)

if ws_all:
    mean_ws = sum(ws_all) / len(ws_all)
    max_ws  = max(ws_all)
    print(f"\n  pan2 WS (QC=1): n={len(ws_all)}, середнє={mean_ws:.2f} м/с, max={max_ws:.1f} м/с")
    print(f"\n  Середня WS по місяцях (QC=1):")
    for mm in range(1, 13):
        if ws_by_month[mm]:
            m_mean = sum(ws_by_month[mm]) / len(ws_by_month[mm])
            print(f"    {MONTHS_UK[mm]:10}: {m_mean:.2f} м/с  (n={len(ws_by_month[mm])})")
        else:
            print(f"    {MONTHS_UK[mm]:10}: — немає даних")
    print(f"\n  ➡ У дисертації: «середня швидкість вітру 2,5 м/с»")
    print(f"     Реальне значення: {mean_ws:.2f} м/с")
    if abs(mean_ws - 2.5) < 0.1:
        print("     ✓ ПІДТВЕРДЖЕНО")
    else:
        print(f"     ✗ РОЗБІЖНІСТЬ — потрібно виправити у тексті на {mean_ws:.1f} м/с")
else:
    print("  Дані WS відсутні!")


# ══════════════════════════════════════════════════════════════
# Q3/Q6 — СТАТИСТИКА ПО ДАТЧИКАХ (pan2, pan5, pan6)
# ══════════════════════════════════════════════════════════════
print("\n" + SEP)
print("  Q3/Q6 — СТАТИСТИКА ПО ДАТЧИКАХ (підрозділ 4.4.2)")
print(SEP)

pan5_data = load_caircloud(os.path.join(DIG_CC, 'CairCloud_DIG_pan5'), has_ws=False)
pan6_data = load_caircloud(os.path.join(DIG_CC, 'CairCloud_DIG_pan6'), has_ws=False)

sensors = [
    ('pan2', pan2_data, True),
    ('pan5', pan5_data, False),
    ('pan6', pan6_data, False),
]

for name, data, has_ws in sensors:
    pm10_qc1 = []
    pm10_all = []
    total_rows = 0
    qc1_rows   = 0
    for mm in range(1, 13):
        for r in data[mm]:
            total_rows += 1
            if r['pm10'] is not None and r['pm10'] >= 0:
                pm10_all.append(r['pm10'])
            if r['qc'] == 1:
                qc1_rows += 1
                if r['pm10'] is not None and r['pm10'] >= 0:
                    pm10_qc1.append(r['pm10'])

    if pm10_qc1:
        mean_pm10 = sum(pm10_qc1) / len(pm10_qc1)
        max_pm10  = max(pm10_qc1)
        pct_qc1   = 100 * qc1_rows / total_rows if total_rows else 0
        print(f"\n  {name}:")
        print(f"    Всього рядків: {total_rows}")
        print(f"    QC=1 рядків:  {qc1_rows} ({pct_qc1:.1f}%)")
        print(f"    PM₁₀ (QC=1):  n={len(pm10_qc1)}, середнє={mean_pm10:.2f} мкг/м³, max={max_pm10:.1f} мкг/м³")
    else:
        print(f"\n  {name}: дані відсутні!")

# Дисертаційні твердження
print(f"\n  ➡ У дисертації:")
print(f"     pan2: річне серед. 26.1 мкг/м³, max 360 мкг/м³, 29042 рядки QC=1 (82.9%)")
print(f"     pan5: річне серед. 21.4 мкг/м³, max 45 мкг/м³,  26043 рядки QC=1 (74.3%)")
print(f"     pan6: річне серед. 25.25 мкг/м³, max 366.6 мкг/м³, 29977 рядки QC=1 (85.5%)")


# ══════════════════════════════════════════════════════════════
# Q4 — QC РЕЗУЛЬТАТИ TZA_Post7
# ══════════════════════════════════════════════════════════════
print("\n" + SEP)
print("  Q4 — QC РЕЗУЛЬТАТИ TZA_Post7 (621/48/573)")
print(SEP)

tza7_all = load_tza_full(os.path.join(DIG_TZA, 'TZA_DIG_Post7'))
tza3_all = load_tza_full(os.path.join(DIG_TZA, 'TZA_DIG_Post3'))

for label, rows_all in [('TZA_Post3', tza3_all), ('TZA_Post7', tza7_all)]:
    total = 0
    missing_tsp = 0
    valid = 0
    high_tsp = 0  # TSP > 800 мкг/м³ (типовий поріг відхилення)
    strok_data = []
    for (mm, row) in rows_all:
        if len(row) < 2:
            continue
        try:
            strok = int(float(str(row[1]).replace(',', '.')))
        except:
            continue
        if strok not in (7, 19):
            continue
        total += 1
        tsp_val = row[2]
        if tsp_val is None:
            missing_tsp += 1
            continue
        try:
            tsp_f = float(str(tsp_val).replace(',', '.'))
            if tsp_f > 800:
                high_tsp += 1
            valid += 1
            strok_data.append(tsp_f)
        except:
            missing_tsp += 1

    print(f"\n  {label}:")
    print(f"    Всього строків 7+19: {total}")
    print(f"    TSP відсутній (None): {missing_tsp}")
    print(f"    TSP присутній: {valid}")
    print(f"    TSP > 800 мкг/м³: {high_tsp}")
    if strok_data:
        print(f"    TSP середнє: {sum(strok_data)/len(strok_data):.1f} мкг/м³")
        print(f"    TSP max: {max(strok_data):.1f} мкг/м³")
        vals_high = sorted([v for v in strok_data if v > 300], reverse=True)[:5]
        if vals_high:
            print(f"    TSP > 300: {vals_high}")

print(f"\n  ➡ У дисертації: Post7: 621 вимірювань, 48 відсутні, 573 пройшли QC")
print(f"     Post3: 624 значення TSP, 100% повнота")


# ══════════════════════════════════════════════════════════════
# Q5 — ТАБЛИЦЯ 4.6: ПОВНОТА РЯДУ pan5 ПО МІСЯЦЯХ
# ══════════════════════════════════════════════════════════════
print("\n" + SEP)
print("  Q5 — ТАБЛИЦЯ 4.6: ПОВНОТА РЯДУ pan5 ПО МІСЯЦЯХ")
print(SEP)

# Очікувана кількість записів на місяць: 2 на годину × 24 год × N днів
import calendar
year = 2025
print(f"\n  {'Місяць':10} {'QC=1':>8} {'Всього':>8} {'Очік.':>8} {'% QC=1/очік':>13} {'% QC=1/всього':>14}")
print(f"  {'-'*10} {'-'*8} {'-'*8} {'-'*8} {'-'*13} {'-'*14}")
total_qc1 = 0
total_all = 0
total_exp = 0
for mm in range(1, 13):
    days = calendar.monthrange(year, mm)[1]
    expected = days * 24 * 2  # кожні 30 хв
    qc1_cnt = sum(1 for r in pan5_data[mm] if r['qc'] == 1)
    all_cnt  = len(pan5_data[mm])
    pct_exp  = 100 * qc1_cnt / expected if expected else 0
    pct_all  = 100 * qc1_cnt / all_cnt  if all_cnt  else 0
    total_qc1 += qc1_cnt
    total_all += all_cnt
    total_exp += expected
    print(f"  {MONTHS_UK[mm]:10} {qc1_cnt:>8} {all_cnt:>8} {expected:>8} {pct_exp:>12.1f}% {pct_all:>13.1f}%")
pct_tot_exp = 100 * total_qc1 / total_exp if total_exp else 0
pct_tot_all = 100 * total_qc1 / total_all if total_all else 0
print(f"  {'РАЗОМ':10} {total_qc1:>8} {total_all:>8} {total_exp:>8} {pct_tot_exp:>12.1f}% {pct_tot_all:>13.1f}%")

print(f"\n  ➡ У дисертації (Таблиця 4.6): 26043 рядки QC=1 (74.3%)")


# ══════════════════════════════════════════════════════════════
# Q7 — ТАБЛИЦЯ 4.7: ПОПАРНІ ВІДНОШЕННЯ МІСЯЧНИХ СЕРЕДНІХ PM₁₀
# ══════════════════════════════════════════════════════════════
print("\n" + SEP)
print("  Q7 — ТАБЛИЦЯ 4.7: МІСЯЧНІ СЕРЕДНІ PM₁₀ І ВІДНОШЕННЯ")
print(SEP)

def monthly_means(data):
    """Обчислити середні PM₁₀ по місяцях (тільки QC=1)."""
    means = {}
    for mm in range(1, 13):
        vals = [r['pm10'] for r in data[mm] if r['qc'] == 1
                and r['pm10'] is not None and r['pm10'] >= 0]
        means[mm] = sum(vals) / len(vals) if vals else None
    return means

m2 = monthly_means(pan2_data)
m5 = monthly_means(pan5_data)
m6 = monthly_means(pan6_data)

print(f"\n  {'Місяць':10} {'pan2':>8} {'pan5':>8} {'pan6':>8} {'pan6/pan2':>10} {'pan2/pan5':>10}")
print(f"  {'-'*10} {'-'*8} {'-'*8} {'-'*8} {'-'*10} {'-'*10}")
for mm in range(1, 13):
    v2 = m2[mm]; v5 = m5[mm]; v6 = m6[mm]
    r62 = round(v6 / v2, 2) if (v2 and v6) else None
    r25 = round(v2 / v5, 2) if (v2 and v5) else None
    s2  = f"{v2:.2f}" if v2 else "—"
    s5  = f"{v5:.2f}" if v5 else "—"
    s6  = f"{v6:.2f}" if v6 else "—"
    sr62 = f"{r62:.2f}" if r62 else "—"
    sr25 = f"{r25:.2f}" if r25 else "—"
    print(f"  {MONTHS_UK[mm]:10} {s2:>8} {s5:>8} {s6:>8} {sr62:>10} {sr25:>10}")

# Річні
all_means = {}
for name_s, data_s in [('pan2', pan2_data), ('pan5', pan5_data), ('pan6', pan6_data)]:
    vals = [r['pm10'] for mm in range(1,13) for r in data_s[mm]
            if r['qc'] == 1 and r['pm10'] is not None and r['pm10'] >= 0]
    all_means[name_s] = sum(vals) / len(vals) if vals else None
    print(f"\n  {name_s}: річне середнє PM₁₀ (QC=1) = {all_means[name_s]:.2f} мкг/м³" if all_means[name_s] else f"\n  {name_s}: —")

# ══════════════════════════════════════════════════════════════
# ПІДСУМОК ВЕРИФІКАЦІЇ
# ══════════════════════════════════════════════════════════════
print("\n" + SEP)
print("  ПІДСУМОК: ВІДПОВІДНІСТЬ ТЕКСТУ ДИСЕРТАЦІЇ РЕАЛЬНИМ ДАНИМ")
print(SEP)
print("""
  Q2 — Теоретична похибка усереднення <5%:
     Це аналітична оцінка. Для вікна ±30 хв і кроку вимірювань 30 хв
     CairCloud виконує до 2 вимірювань на вікно (n=1-2). При σ/μ ~30%
     (типовий CV для PM₁₀) і n=2: похибка середнього = σ/√n = 0.3/√2 = 21%.
     Значення "<5%" не підтверджується простим формульним розрахунком.
     РЕКОМЕНДАЦІЯ: замінити на конкретне число з даних або посилання
     на методичну літературу (ДСТУ, РД 52).
""")

print(SEP)
print("  ФАЙЛ ВЕРИФІКАЦІЇ ВИКОНАНО")
print(SEP)
