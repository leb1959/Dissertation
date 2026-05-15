# -*- coding: utf-8 -*-
"""
СКРИПТ ЗБІРКИ ТАБЛИЦІ 4.8
Зведена статистика верифікованого масиву ТЗА
(TZA_Post3 та TZA_Post7, строки 7 та 19, TSP≠None)
Джерело: Processed_TZA_DIG/
Вихід:   Таблиця_4_8.xlsx
"""
import os, sys
from zipfile import BadZipFile
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TZA_DIR    = os.path.join(SCRIPT_DIR, "Додаток_2 Оцифровані дані", "Processed_TZA_DIG")
OUT_FILE   = os.path.join(SCRIPT_DIR, "Таблиця_4_8.xlsx")
MONTHS_UK  = ['','Січень','Лютий','Березень','Квітень','Травень','Червень',
               'Липень','Серпень','Вересень','Жовтень','Листопад','Грудень']
POSTS      = [('TZA_Post3','TZA_DIG_Post3'),('TZA_Post7','TZA_DIG_Post7')]
# Колонки у DIG: День(0) Строк(1) TSP(2) T(3) WD(4) WS(5) ATM(6)
TSP_COL, STR_COL = 2, 1
VALID_STROK = {7, 19}

def load_post(post_dir):
    monthly = {}
    folder  = os.path.basename(post_dir)
    for mm in range(1,13):
        fpath = os.path.join(post_dir, f"{mm:02d}_{folder}.xlsx")
        monthly[mm] = {'n':0,'vals':[]}
        if not os.path.exists(fpath): continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [!] {os.path.basename(fpath)}: {e}"); continue
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if len(row) <= TSP_COL: continue
            if row[STR_COL] not in VALID_STROK: continue
            tsp = row[TSP_COL]
            if tsp is None: continue
            try:
                v = float(tsp)
                if v >= 0:
                    monthly[mm]['n']   += 1
                    monthly[mm]['vals'].append(v)
            except: pass
        wb.close()
    return monthly

print("="*60)
print("  ЗБІРКА ТАБЛИЦІ 4.8 — Зведена статистика ТЗА")
print("="*60)

all_data = {}
for post_name, subfolder in POSTS:
    print(f"\n  [{post_name}]")
    post_path = os.path.join(TZA_DIR, subfolder)
    data = load_post(post_path)
    all_data[post_name] = data
    for mm in range(1,13):
        n = data[mm]['n']
        mean = round(sum(data[mm]['vals'])/n,0) if n else None
        print(f"    {MONTHS_UK[mm]:10}: n={n:4}  mean={mean}")

# Стилі
thin   = Side(style='thin')
border = Border(left=thin,right=thin,top=thin,bottom=thin)
hfill  = PatternFill('solid', fgColor='D9E1F2')
tfont  = Font(name='Times New Roman', bold=True, size=11)
hfont  = Font(name='Times New Roman', bold=True, size=10)
dfont  = Font(name='Times New Roman', size=10)
bfont  = Font(name='Times New Roman', bold=True, size=10)
center = Alignment(horizontal='center',vertical='center',wrap_text=True)
left   = Alignment(horizontal='left',vertical='center')

wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = "Таблиця 4.8"

ws.merge_cells('A1:E1')
ws['A1'] = 'Таблиця 4.8. Зведена статистика верифікованого масиву ТЗА (2025 р., Кривий Ріг)'
ws['A1'].font = tfont; ws['A1'].alignment = center
ws.row_dimensions[1].height = 26

HEADERS = ['Місяць','TZA_Post3\n(вимірювань)','Середнє TSP\nPost3, мкг/м³',
           'TZA_Post7\n(вимірювань)','Середнє TSP\nPost7, мкг/м³']
WIDTHS  = [12,16,16,16,16]
for ci,(h,w) in enumerate(zip(HEADERS,WIDTHS),1):
    c = ws.cell(2,ci,h)
    c.font=hfont; c.alignment=center; c.border=border; c.fill=hfill
    ws.column_dimensions[get_column_letter(ci)].width=w
ws.row_dimensions[2].height=34

tot3_n, tot3_v = 0, []
tot7_n, tot7_v = 0, []
for ri,mm in enumerate(range(1,13),3):
    d3 = all_data['TZA_Post3'][mm]
    d7 = all_data['TZA_Post7'][mm]
    n3 = d3['n']; m3 = round(sum(d3['vals'])/n3,0) if n3 else None
    n7 = d7['n']; m7 = round(sum(d7['vals'])/n7,0) if n7 else None
    tot3_n+=n3; tot3_v+=d3['vals']
    tot7_n+=n7; tot7_v+=d7['vals']
    for ci,v in enumerate([MONTHS_UK[mm],n3,int(m3) if m3 else '—',n7,int(m7) if m7 else '—'],1):
        c=ws.cell(ri,ci,v); c.font=dfont; c.border=border
        c.alignment=left if ci==1 else center

# Разом
m3t = int(round(sum(tot3_v)/tot3_n,0)) if tot3_n else '—'
m7t = int(round(sum(tot7_v)/tot7_n,0)) if tot7_n else '—'
for ci,v in enumerate(['Разом / Сер.',tot3_n,m3t,tot7_n,m7t],1):
    c=ws.cell(15,ci,v); c.font=bfont; c.border=border
    c.alignment=left if ci==1 else center
    c.fill=PatternFill('solid',fgColor='F2F2F2')

ws.merge_cells('A16:E16')
ws['A16'] = ('Примітка. Враховано строки 7 та 19 год з TSP≠None. '
             'Джерело: Processed_TZA_DIG/ (build_Таблиця_4_8.py).')
ws['A16'].font = Font(name='Times New Roman',italic=True,size=9)
ws['A16'].alignment = Alignment(wrap_text=True,vertical='top')
ws.row_dimensions[16].height=35
ws.freeze_panes='A3'
wb_out.save(OUT_FILE)
print(f"\n  ✓ ТАБЛИЦЯ 4.8: {OUT_FILE}")
print(f"  Post3: {tot3_n} вим., сер.={m3t} мкг/м³")
print(f"  Post7: {tot7_n} вим., сер.={m7t} мкг/м³")
