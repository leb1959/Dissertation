# -*- coding: utf-8 -*-
"""
СКРИПТ ЗБІРКИ ТАБЛИЦІ 4.12
Основні метеорологічні параметри, 2025 р. (Кривий Ріг)
  T, WS  — з TZA_DIG_Post3 (строки 7 та 19; надійний метеощогловий вимір)
  RH, P  — з CairCloud_DIG_pan2 (записи QC=1)
  Обґрунтування: внутрішній термометр та анемометр Cairnet pan2 дають
  фізично неправдоподібні значення (T=+24°C у листопаді, WS≈0.2 м/с),
  тому для T і WS використовується надійний TZA_Post3.
Вихід: Таблиця_4_12.xlsx
"""
import os, sys
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TZA3_DIR   = os.path.join(SCRIPT_DIR, "Додаток_2 Оцифровані дані",
                           "Processed_TZA_DIG", "TZA_DIG_Post3")
PAN2_DIR   = os.path.join(SCRIPT_DIR, "Додаток_2 Оцифровані дані",
                           "Processed_CairCloud_DIG", "CairCloud_DIG_pan2")
OUT_FILE   = os.path.join(SCRIPT_DIR, "Таблиця_4_12.xlsx")

MONTHS_UK = ['','Січень','Лютий','Березень','Квітень','Травень','Червень',
             'Липень','Серпень','Вересень','Жовтень','Листопад','Грудень']
VALID_STROK = {7, 19}

# TZA: День(0) Строк(1) TSP(2) T(3) WD(4) WS(5) ATM(6)
# pan2: Timestamp(0) PM10(1) QC(2) T(3) QC_T(4) RH(5) QC_RH(6) PA(7) QC_PA(8) WS(9)

print("="*60)
print("  ЗБІРКА ТАБЛИЦІ 4.12 — Метеорологічні параметри")
print("  T,WS: TZA_Post3 | RH,P: CairCloud pan2 (QC=1)")
print("="*60)

rows_data = []
yr_t, yr_ws, yr_rh, yr_pa = [], [], [], []

for mm in range(1, 13):
    # --- T, WS з TZA Post3 ---
    tza_path = os.path.join(TZA3_DIR, f"{mm:02d}_TZA_DIG_Post3.xlsx")
    t_vals, ws_vals = [], []
    if os.path.exists(tza_path):
        try:
            wb = openpyxl.load_workbook(tza_path, read_only=True, data_only=True)
            for row in wb.active.iter_rows(min_row=2, values_only=True):
                if not row or row[1] not in VALID_STROK: continue
                try:
                    if row[3] is not None: t_vals.append(float(row[3]))
                    if row[5] is not None: ws_vals.append(float(row[5]))
                except Exception: pass
            wb.close()
        except Exception as e:
            print(f"  [!] TZA {mm:02d}: {e}")
    else:
        print(f"  [!] TZA файл не знайдено: {tza_path}")

    # --- RH, P з CairCloud pan2 (QC=1) ---
    cc_path = os.path.join(PAN2_DIR, f"{mm:02d}_CairCloud_DIG_pan2.xlsx")
    rh_vals, pa_vals = [], []
    if os.path.exists(cc_path):
        try:
            wb = openpyxl.load_workbook(cc_path, read_only=True, data_only=True)
            for row in wb.active.iter_rows(min_row=2, values_only=True):
                if not row or row[2] != 1: continue
                try:
                    if row[5] is not None: rh_vals.append(float(row[5]))
                    if row[7] is not None: pa_vals.append(float(row[7]))
                except Exception: pass
            wb.close()
        except Exception as e:
            print(f"  [!] pan2 {mm:02d}: {e}")
    else:
        print(f"  [!] pan2 файл не знайдено: {cc_path}")

    mean_t  = round(sum(t_vals)/len(t_vals),   1) if t_vals  else None
    mean_ws = round(sum(ws_vals)/len(ws_vals), 1) if ws_vals else None
    mean_rh = int(round(sum(rh_vals)/len(rh_vals), 0)) if rh_vals else None
    mean_pa = int(round(sum(pa_vals)/len(pa_vals), 0)) if pa_vals else None

    yr_t  += t_vals;  yr_ws += ws_vals
    yr_rh += rh_vals; yr_pa += pa_vals

    rows_data.append((MONTHS_UK[mm], mean_t, mean_rh, mean_ws, mean_pa))
    print(f"  {MONTHS_UK[mm]:10}: T={mean_t}°C  RH={mean_rh}%  "
          f"WS={mean_ws}м/с  P={mean_pa}гПа")

yr_T  = round(sum(yr_t)/len(yr_t),   1) if yr_t  else None
yr_RH = int(round(sum(yr_rh)/len(yr_rh), 0)) if yr_rh else None
yr_WS = round(sum(yr_ws)/len(yr_ws), 1) if yr_ws else None
yr_PA = int(round(sum(yr_pa)/len(yr_pa), 0)) if yr_pa else None

# ── Запис xlsx ────────────────────────────────────────────────────
thin   = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill  = PatternFill('solid', fgColor='D9E1F2')
tfont  = Font(name='Times New Roman', bold=True, size=11)
hfont  = Font(name='Times New Roman', bold=True, size=10)
dfont  = Font(name='Times New Roman', size=10)
bfont  = Font(name='Times New Roman', bold=True, size=10)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left',   vertical='center')

wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = "Таблиця 4.12"

ws.merge_cells('A1:E1')
ws['A1'] = ('Таблиця 4.12. Основні метеорологічні параметри (2025 р., Кривий Ріг). '
            'T, WS — TZA_Post3 (строки 7 та 19); RH, P — Cairnet pan2 (QC=1).')
ws['A1'].font = tfont; ws['A1'].alignment = center
ws.row_dimensions[1].height = 32

HEADERS = ['Місяць', 'Сер. Т,\n°C', 'Сер. RH,\n%', 'Сер. WS,\nм/с', 'Сер. P,\nгПа']
WIDTHS  = [12, 10, 10, 10, 10]
for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
    c = ws.cell(2, ci, h)
    c.font = hfont; c.alignment = center; c.border = border; c.fill = hfill
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[2].height = 34

for ri, (month, t, rh, ws_v, pa) in enumerate(rows_data, 3):
    for ci, v in enumerate([month,
                             t   if t   is not None else '—',
                             rh  if rh  is not None else '—',
                             ws_v if ws_v is not None else '—',
                             pa  if pa  is not None else '—'], 1):
        c = ws.cell(ri, ci, v)
        c.font = dfont; c.border = border
        c.alignment = left if ci == 1 else center

for ci, v in enumerate(['Рік/Сер.', yr_T, yr_RH, yr_WS, yr_PA], 1):
    c = ws.cell(15, ci, v if v is not None else '—')
    c.font = bfont; c.border = border
    c.alignment = left if ci == 1 else center
    c.fill = PatternFill('solid', fgColor='F2F2F2')

ws.merge_cells('A16:E16')
ws['A16'] = ('Примiтка. T та WS - середнi за строками 7 та 19 год з TZA_DIG_Post3 '
             '(внутрiшнi датчики Cairnet дають фiзично неправдоподiбнi значення). '
             'RH та P - середнi за записами QC=1 з CairCloud_DIG_pan2. '
             'Джерело: build_Таблиця_4_12.py.')
ws['A16'].font = Font(name='Times New Roman', italic=True, size=9)
ws['A16'].alignment = Alignment(wrap_text=True, vertical='top')
ws.row_dimensions[16].height = 40
ws.freeze_panes = 'A3'

wb_out.save(OUT_FILE)
print('  Таблиця 4.12 збережено: ' + OUT_FILE)
print('  Рiк: T=' + str(yr_T) + 'C  RH=' + str(yr_RH) + '%  WS=' + str(yr_WS) + 'м/с  P=' + str(yr_PA) + 'гПа')
