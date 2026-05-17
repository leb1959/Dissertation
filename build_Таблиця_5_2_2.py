"""
build_Таблиця_5_2_2.py
Генерує Таблиця_5_2_2.xlsx — показники впливу відносної вологості на PM10
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "Таблиця_5_2_2.xlsx"

def hdr(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=True, color="FFFFFF", name='Times New Roman', size=11)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def dat(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name='Times New Roman', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Таблиця 5.2.2"

ws.merge_cells('A1:D1')
ws['A1'].value = "Таблиця 5.2.2 – Узагальнені показники для аналізу впливу зовнішньої відносної вологості"
ws['A1'].font = Font(bold=True, name='Times New Roman', size=12)
ws['A1'].alignment = Alignment(horizontal='center', wrap_text=True)
ws.row_dimensions[1].height = 30

for c, h in enumerate(['Показник', 'POST2', 'POST6', 'pan5'], 1):
    hdr(ws, 2, c, h)
ws.row_dimensions[2].height = 35

RH_DATA = [
    ('r(PM₁₀–RH), річна вибірка', 0.36, 0.42, 0.45),
    ('Міжпостова узгодженість PM₁₀ (холодний період)*', 0.69, 0.66, 0.72),
    ('Міжпостова узгодженість PM₁₀ (теплий період)*', 0.55, 0.45, 0.49),
]
for r, row in enumerate(RH_DATA, 3):
    dat(ws, r, 1, row[0])
    for c in range(1, 4):
        dat(ws, r, c+1, row[c])

ws.merge_cells('A6:D6')
ws['A6'].value = ("* У другому та третьому рядках наведено сезонні міжпостові коефіцієнти кореляції PM₁₀, "
                  "що характеризують загальний рівень узгодженості вимірювань між усіма трьома парами постів.")
ws['A6'].font = Font(name='Times New Roman', size=10, italic=True)
ws['A6'].alignment = Alignment(wrap_text=True)
ws.row_dimensions[6].height = 40

for i, w in enumerate([48, 10, 10, 10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print(f"Збережено: {OUT}")
