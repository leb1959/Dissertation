"""
build_Таблиця_5_2_3.py
Генерує Таблиця_5_2_3.xlsx — показники впливу температури повітря на PM10
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "Таблиця_5_2_3.xlsx"

def hdr(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=True, color="FFFFFF", name='Times New Roman', size=11)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def dat(ws, r, c, v, fill=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name='Times New Roman', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Таблиця 5.2.3"

ws.merge_cells('A1:D1')
ws['A1'].value = "Таблиця 5.2.3 – Узагальнені показники для аналізу впливу зовнішньої температури"
ws['A1'].font = Font(bold=True, name='Times New Roman', size=12)
ws['A1'].alignment = Alignment(horizontal='center', wrap_text=True)
ws.row_dimensions[1].height = 30

for c, h in enumerate(['Показник', 'POST2', 'POST6', 'pan5'], 1):
    hdr(ws, 2, c, h)
ws.row_dimensions[2].height = 35

T_DATA = [
    ('Міжпостова узгодженість PM₁₀ (річна вибірка)*', 0.74, 0.72, 0.77),
    ('r(PM₁₀–T), холодний період', -0.36, -0.30, -0.21),
    ('Найбільш виражений місячний ефект', 'лютий', 'лютий', 'жовтень'),
]
for r, row in enumerate(T_DATA, 3):
    dat(ws, r, 1, row[0])
    for c in range(1, 4):
        fill = "FFEB9C" if isinstance(row[c], float) and row[c] <= -0.30 else None
        dat(ws, r, c+1, row[c], fill=fill)

ws.merge_cells('A6:D6')
ws['A6'].value = ("* У першому рядку наведено міжпостові коефіцієнти кореляції PM₁₀ для повної спільної вибірки, "
                  "що характеризують загальну міжпостову узгодженість даних.")
ws['A6'].font = Font(name='Times New Roman', size=10, italic=True)
ws['A6'].alignment = Alignment(wrap_text=True)
ws.row_dimensions[6].height = 40

for i, w in enumerate([48, 12, 12, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print(f"Збережено: {OUT}")
